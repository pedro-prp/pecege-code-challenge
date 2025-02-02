from openpyxl import load_workbook, Workbook
from io import BytesIO

from .repositories import PersonRepository
from .serializers import PersonSerializer

from .utils import calculate_age, calculate_value

from datetime import datetime


class PersonService:
    def __init__(self):
        self.__repository = PersonRepository()

    def process_spreadsheet(self, file):
        workbook = load_workbook(file)
        worksheet = workbook.active

        processed_data = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None for cell in row):
                continue

            try:
                name, email, birth_date, is_active, _ = row

                if isinstance(birth_date, datetime):
                    birth_date = birth_date.date()

                age = calculate_age(birth_date)

                if age < 18:
                    is_active = False

                value = calculate_value(age)

                person = {
                    "name": name,
                    "email": email,
                    "birth_date": birth_date,
                    "is_active": is_active,
                    "value": value,
                }

                serializer = PersonSerializer(data=person)

                if serializer.is_valid():
                    self.__repository.update_or_create(serializer.validated_data)

                if is_active:
                    processed_data.append(person)

            except Exception as e:
                print(f"Erro ao processar a linha: {row}. Error: {e}")
                continue

        return processed_data

    def generate_spreadsheet(self):
        data = self.__repository.get_all()

        workbook = Workbook()
        worksheet = workbook.active

        worksheet.title = "Pessoas"

        column_names = [
            "Nome",
            "Email",
            "Data de Nascimento",
            "Ativo",
            "Valor",
        ]

        worksheet.append(column_names)

        for person in data:
            worksheet.append(
                [
                    person["name"],
                    person["email"],
                    person["birth_date"].strftime("%d/%m/%Y"),
                    person["is_active"],
                    float(person["value"]),
                ]
            )

        output_file = BytesIO()
        workbook.save(output_file)

        output_file.seek(0)

        return output_file
